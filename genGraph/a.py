import networkx as nx
import matplotlib.pyplot as plt
import datetime
import random
import time
import numpy as np # type: ignore
import math
import sys
import openpyxl.workbook
import pandas as pd
import os
import xlwt , xlutils
from xlwt import Workbook
from xlrd import open_workbook
# from dataForm import *

import openpyxl

def genBiGraph( _numL , _numR , _averageDegreeR,_seed):
    
    workPath = r'./'
    execFileName = 'log.xlsx'
    dataSetDirName = '/data/'
    

    execFilePath = workPath + execFileName  
    dataSetDirPath = workPath+dataSetDirName 

    if(_seed == -1):
        seeds = int(time.time())
        np.random.seed(seeds)
    else:
        seeds = _seed
        np.random.seed(seeds)

    numL = _numL
    numR = _numR
    maxDegreeR = _averageDegreeR * 2

    degreeR = np.random.uniform(0,1,numR)


    nodeL = [x for x in range(0,numL,1)]

    degreeR=[ min(math.ceil(x*maxDegreeR),numL) for x in degreeR ]
      
    numE = sum(degreeR) 

    # seed.NL.NR.NE.time
    dateTime = str(datetime.datetime.now())
    dateTime=dateTime[:dateTime.rfind(":",1)]
    dateTime=dateTime.replace(" ","-")
    dateTime=dateTime.replace(":","")
    fileName =  dateTime + "---%d.%d.%d.%d"%(numL,numR ,numE,seeds)
    print(fileName)

    filePath = f"{dataSetDirPath}//{fileName}//"
    if not os.path.exists(filePath):
        os.makedirs(filePath)

    degreeL = [ 0 for x in range(numL) ]

    with open(filePath+"originG","w") as file :
        time.sleep(0.1)
        for index in range(0,numR) : #nodeR在这里产生的
            for L in random.sample(nodeL,degreeR[index]):
                degreeL[L]+=1
                file.writelines("%d     %d\n"%(L,index))
        file.close
    
    #新增一行
    file = open(filePath+"originG","r+")
    lines = file.readlines()
    file.seek(0)
    file.write("# %d %d %d\n"%(numL,numR,sum(degreeR)))
    file.writelines(lines)


    graphInfo = [
        ['name' , str(fileName)],
        ['numL' , numL ],
        ['numR' , numR ], 
        ['numE' , sum(degreeR) ], 
        ['averageDegree' , (numE *2) / (numL+numR)],
        ['averageDegreeL' , numE / numL],
        ['averageDegreeR' , numE / numR],
        ['maxDegreeL' , max(degreeL) ],
        ['maxDegreeR' , max(degreeR) ]
    ]

    if not os.path.exists(execFilePath) :
        xls = openpyxl.Workbook()
        graphSheet = xls.active
        graphSheet.title='dataset'
        for colIndex in range(len(graphInfo)) :
            print(graphInfo[colIndex][1])
            graphSheet.cell(1,colIndex+1,value=graphInfo[colIndex][0])
        xls.save(execFilePath)
    

    xls = openpyxl.load_workbook(execFilePath)
    graphSheet = xls['dataset']
    row_num = graphSheet.max_row
    for colIndex in range(len(graphInfo)) :
        graphSheet.cell(row_num + 1 , colIndex+1 , graphInfo[colIndex][1])

    
    xls.save(execFilePath)



if __name__ == "__main__":

     
    numL = int(sys.argv[1])
    numR = int(sys.argv[2])
    averageDegreeR = int(sys.argv[3]) 
    if(len(sys.argv) < 4):
        genBiGraph(numL,numR,averageDegreeR,-1)
    else:
        genBiGraph(numL,numR,averageDegreeR,int(sys.argv[4]))




    


        
        





